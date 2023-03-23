import datetime
import logging
import re
from configparser import ConfigParser
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.cell import Cell, ReadOnlyCell
from openpyxl.cell.read_only import EmptyCell

from vfl2csv_base.ColumnScheme import ColumnScheme

TrialSiteMetadata = dict[str, str]
TrialSiteHeader = list[tuple[int, str]]
TrialSiteData = list[list[str]]
TrialSiteContent = tuple[TrialSiteMetadata, TrialSiteHeader, TrialSiteData]

ExcelCell = Cell | ReadOnlyCell | EmptyCell


class ConversionAuditor:
    logger = logging.getLogger(__name__)

    def __init__(self, vfl2csv_config: ConfigParser, column_scheme: ColumnScheme):
        self.config = vfl2csv_config
        self.column_scheme = column_scheme
        self.reference: list[TrialSiteContent] = []

    def set_reference_files(self, paths: list[Path]) -> None:
        self.reference = []
        file_type = self.config['Input']['input_format']
        if file_type.lower() == 'excel':
            for path in paths:
                self.reference.extend(self._parse_excel_file(path))
        elif file_type.lower() == 'tsv':
            for path in paths:
                self.reference.append(self._parse_tsv_file(path))
        else:
            raise ValueError('`file_type` must be either "Excel" or "TSV"!')

    def _parse_excel_file(self, file: Path) -> list[TrialSiteContent]:
        wb = load_workbook(file, read_only=True)
        results: list[TrialSiteContent] = []
        for sheet in wb.sheetnames:
            lines: list[list[str]] = list(wb[sheet].rows)

            results.append(self._parse_input(lines))
        return results

    def _parse_tsv_file(self, file: Path) -> TrialSiteContent:
        with open(file, 'r', encoding=self.config['Input']['tsv_encoding']) as file:
            raw_lines = file.readlines()
        lines: list[list[str]] = [line.split('\t') for line in raw_lines]

        # TSV files appear to have one more empty line between metadata and data than the Excel sheets have.
        # Simply delete one of these empty lines.
        del lines[13]

        # remove newlines at the end of cells or with their own cells
        for line in lines:
            if line[-1] == '\n':
                del line[-1]
            elif line[-1].endswith('\n'):
                line[-1] = line[-1][0:-1]

        return self._parse_input(lines)

    def _parse_input(self, lines: list[Sequence[ExcelCell | str]]) -> TrialSiteContent:
        metadata: TrialSiteMetadata = {}
        header: TrialSiteHeader = []
        data: TrialSiteData = []
        for line in lines[4:11]:
            string = line[0].value if isinstance(line[0], ExcelCell) else line[0]
            key, value = string.split(':')
            metadata[key.strip()] = re.sub(r'\s+', ' ', value.strip())

        years = []
        for column in lines[13]:
            value: str | datetime.datetime = column.value if isinstance(column, ExcelCell) else column
            if isinstance(value, datetime.datetime):
                month, year = value.month, value.year
                years.append(year - (0 if month > 5 else 1))
            elif re.fullmatch(r'\d{2}\.\d{2}\.\d{4}', value):
                day, month, year = value.split('.')
                years.append(int(year) - (0 if int(month) > 5 else 1))
            else:
                years.append(-1)
        labels = [column['override_name'] for column in self.column_scheme.head]
        if (len(years) - len(self.column_scheme.head)) % len(self.column_scheme.measurements) != 0:
            raise ValueError('Invalid column count')
        measurement_count = (len(years) - len(self.column_scheme.head)) // len(self.column_scheme.measurements)
        labels.extend(measurement_count * [column['override_name'] for column in self.column_scheme.measurements])
        header.extend(list(zip(years, labels)))
        for line in lines[17:]:
            line_tokens = []
            for cell in line:
                value = cell.value if isinstance(cell, ExcelCell) else cell
                # check for `None` must happen first
                if value is None or value == '':
                    value = 'NA'
                if not isinstance(value, str):
                    value = str(value)
                line_tokens.append(value)
            data.append(line_tokens)
        return metadata, header, data

    def audit_converted_metadata_files(self, paths: list[Path]) -> None:
        reference_sites_index = {}
        for site in self.reference:
            key = (site[0]['Versuch'], site[0]['Parzelle'])
            reference_sites_index[key] = site
        for path in paths:
            self._verify_converted_trial_site(path, reference_sites_index)
        if len(reference_sites_index) != 0:
            remaining_trial_site_names = []
            for key, value in reference_sites_index.items():
                remaining_trial_site_names.append('-'.join(key))
            raise ValueError('Count of remaining reference sites is greater than zero, remaining sites: ' +
                             ', '.join(remaining_trial_site_names))

    @staticmethod
    def _verify_metadata_embedded_path(pattern: str, metadata: dict[str, str], actual_path: Path) \
            -> ValueError | None:
        for key, value in metadata.items():
            pattern = pattern.replace(f'{{{key.lower()}}}', value)
        pattern_tokens = re.split(r'[\/\\]', pattern)
        # get the last N components of the file path (i.e. the filename and the N-1 last directories)
        path_tokens = re.split(r'[\/\\]', str(actual_path.absolute()))[-len(pattern_tokens):]
        if pattern_tokens != path_tokens:
            raise ValueError(f'Actual file path does not match expected file path.\n'
                             f'Actual path: {"/".join(path_tokens)}\n'
                             f'Expected path: {"/".join(pattern_tokens)}')
        return

    def _verify_converted_trial_site(self, path: Path, original_trial_sites: dict[tuple[str, str], TrialSiteContent]):
        metadata: TrialSiteMetadata = {}
        header: TrialSiteHeader = []
        data: TrialSiteData
        # read metadata
        # this implementation is copy-pasted from TrialSite.from_metadata_file
        with open(path, 'r', encoding='utf-8') as file:
            for line in file.readlines():
                # use maxsplit to avoid removing equality symbols in the value
                key, value = line.split('=', maxsplit=1)
                # remove trailing newline in value
                metadata[key] = value.rstrip()

        dataframe_path = path.parent / metadata['DataFrame']
        # delete 'DataFrame' entry because it's not part of the original metadata
        del metadata['DataFrame']
        # check the file path
        self._verify_metadata_embedded_path(self.config['Output']['metadata_output_pattern'], metadata, path)
        self._verify_metadata_embedded_path(self.config['Output']['csv_output_pattern'], metadata, dataframe_path)

        # read header & content
        with open(dataframe_path) as file:
            lines = [line.replace('\n', '').split(',') for line in file.readlines()]
            for cell in lines[0]:
                if re.fullmatch(r'\D+_\d{4}', cell):
                    # measurement column
                    label, year = cell.split('_')
                    header.append((int(year), label))
                else:
                    # tree metadata column
                    header.append((-1, cell))
            data = lines[1:]

        trialsite_key = (metadata['Versuch'], metadata['Parzelle'])
        trialsite_key_str = '-'.join(trialsite_key)
        if trialsite_key not in original_trial_sites:
            raise ValueError(f'[{trialsite_key_str}] No original trial site with matching key')

        original_metadata, original_header, original_data = original_trial_sites[trialsite_key]
        if original_metadata != metadata:
            raise ValueError(f'Metadata of converted trial site {trialsite_key_str} does not match metadata of '
                             f'original trial site')

        if original_header != header:
            raise ValueError(f'Header of converted trial site {trialsite_key_str} does not match header of original '
                             f'trial site.\n'
                             f'Converted: {header}\n'
                             f'Original: {original_header}')

        if len(original_data) != len(data):
            raise ValueError(f'Number of data rows of converted trial site {trialsite_key_str} does not match data '
                             f'of original trial site')

        for row_index in range(len(data)):
            for col_index in range(len(data[row_index])):
                error_flag = False
                if re.fullmatch(r'\d+[.,]\d+', data[row_index][col_index]):
                    data_value = data[row_index][col_index].replace(',', '.')
                    original_data_value = original_data[row_index][col_index].replace(',', '.')
                    if float(data_value) != float(original_data_value):
                        error_flag = True
                else:
                    if data[row_index][col_index] != original_data[row_index][col_index]:
                        error_flag = True
                if error_flag:
                    raise ValueError(f'Data of converted trial site {trialsite_key_str} does not match data of '
                                     f'original trial site in line {row_index + 1}, column {col_index + 1}.\n'
                                     f'Converted: {data[row_index][col_index]}\n'
                                     f'Original: {original_data[row_index][col_index]}')
        del original_trial_sites[trialsite_key]
