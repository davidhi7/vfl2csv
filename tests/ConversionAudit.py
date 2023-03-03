import datetime
import logging
import re
from configparser import ConfigParser
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.cell import Cell, ReadOnlyCell
from openpyxl.cell.read_only import EmptyCell

from vfl2csv_base.ColumnScheme import ColumnScheme

TrialSiteMetadata = dict[str, str]
TrialSiteHeader = list[tuple[int, str]]
TrialSiteData = list[list[str]]
TrialSiteContent = tuple[TrialSiteMetadata, TrialSiteHeader, TrialSiteData]

ExcelCell = Cell | ReadOnlyCell | EmptyCell


class ConversionAudit:
    logger = logging.getLogger(__name__)

    def __init__(self, vfl2csv_config: ConfigParser, column_scheme: ColumnScheme):
        self.config = vfl2csv_config
        self.column_scheme = column_scheme
        self.reference: list[TrialSiteContent] = []

    def set_reference_files(self, paths: list[Path], file_type: str) -> None:
        self.reference = []
        if file_type.lower() == 'excel':
            for path in paths:
                self.reference.extend(self._parse_excel_file(path))
        elif file_type.lower() == 'tsv':
            for path in paths:
                self.reference.extend(self._parse_excel_file(path))

        else:
            raise ValueError('`file_type` must be either "Excel" or "TSV"!')

    def _parse_excel_file(self, file: Path) -> list[TrialSiteContent]:
        wb = load_workbook(file, read_only=True)
        results: list[TrialSiteContent] = []
        for sheet in wb.sheetnames:
            results.append(self._parse_input(list(wb[sheet].rows)))
        return results

    def _parse_tsv_file(self, file: Path) -> TrialSiteContent:
        with open(file, 'r', encoding=self.config['Input']['tsv_encoding']) as file:
            raw_lines = file.readlines()
        lines: list[list[str]] = [line.split('\t') for line in raw_lines]

        del lines[13]

        # remove newlines at the end of cells or with their own cells
        for line in lines:
            if line[-1] == '\n':
                del line[-1]
            elif line[-1].endswith('\n'):
                line[-1] = line[-1][0:-2]

        return self._parse_input(lines)

    def _parse_input(self, lines: list[Sequence[ExcelCell | str]]) -> TrialSiteContent:
        metadata: TrialSiteMetadata = {}
        header: TrialSiteHeader = []
        data: TrialSiteData = []
        for line in lines[4:11]:
            string = line[0].value if isinstance(line[0], ExcelCell) else line[0]
            key, value = string.split(':')
            metadata[key.strip()] = re.sub(r'\s+', ' ', value.strip())

        years = []
        for column in lines[13]:
            value = column.value if isinstance(column, ExcelCell) else column
            # if re.fullmatch(r'\d{2}\.\d{2}\.\d{4}', value):
            if isinstance(value, datetime.datetime):
                # day, month, year = int(value.split('.'))
                # years.append(int(year) - 0 if int(month) > 5 else 1)
                month, year = value.month, value.year
                years.append(year - (0 if month > 5 else 1))
            else:
                years.append(-1)
        labels = [column['override_name'] for column in self.column_scheme.head]
        if (len(years) - len(self.column_scheme.head)) % len(self.column_scheme.measurements) != 0:
            raise ValueError('Invalid column count')
        measurement_count = (len(years) - len(self.column_scheme.head)) // len(self.column_scheme.measurements)
        labels.extend(measurement_count * [column['override_name'] for column in self.column_scheme.measurements])
        header.extend(list(zip(years, labels)))
        for line in lines[17:]:
            line_tokens = []
            for cell in line:
                value = cell.value if isinstance(cell, ExcelCell) else cell
                if value is None:
                    value = 'NA'
                line_tokens.append(value)
            data.append(line_tokens)
        return metadata, header, data

    def audit_converted_metadata_files(self, paths: list[Path]) -> None:
        reference_sites_index = {}
        for site in self.reference:
            key = (site[0]['Versuch'], site[0]['Parzelle'])
            reference_sites_index[key] = site
        for path in paths:
            # TODO error handling
            self._verify_converted_trial_site(path, reference_sites_index)
        if len(reference_sites_index) != 0:
            raise ValueError('Count of remaining reference sites is greater than zero')

    @staticmethod
    def _verify_converted_trial_site(path: Path, original_trial_sites: dict[tuple[str, str], TrialSiteContent]):
        metadata: TrialSiteMetadata = {}
        header: TrialSiteHeader = []
        data: TrialSiteData
        # this implementation is copy-pasted from TrialSite.from_metadata_file
        with open(path, 'r', encoding='utf-8') as file:
            for line in file.readlines():
                # use maxsplit to avoid removing equality symbols in the value
                key, value = line.split('=', maxsplit=1)
                # remove trailing newline in value
                metadata[key] = value.rstrip()
        with open(path.parent / metadata['DataFrame']) as file:
            lines = [line.replace('\n', '').split(',') for line in file.readlines()]
            for cell in lines[0]:
                if re.fullmatch(r'\D+_\d{4}', cell):
                    # measurement column
                    label, year = cell.split('_')
                    header.append((int(year), label))
                else:
                    # tree metadata column
                    header.append((-1, cell))
            data = lines[1:]
            for line in data:
                for i, value in enumerate(line):
                    # match integers (N) or floats with a decimal part of zero (N.0 / N.00)
                    if re.fullmatch(r'\d+(\.0+)?', value):
                        # >>> int(10.0) raises an error; this is solved by calling float() first
                        line[i] = int(float(value))
                    elif re.fullmatch(r'\d+\.\d+', value):
                        line[i] = float(value)

        # delete 'DataFrame' entry because it's not part of the original metadata
        del metadata['DataFrame']

        trialsite_key = (metadata['Versuch'], metadata['Parzelle'])
        if trialsite_key not in original_trial_sites:
            raise ValueError('No original trial site with matching key')

        original_metadata, original_header, original_data = original_trial_sites[trialsite_key]
        if original_metadata != metadata:
            raise ValueError('Metadata of original trial site does not match metadata of original trial site')
        if original_header != header:
            raise ValueError('Header of original trial site does not match header of original trial site')
        if original_data != data:
            raise ValueError('Data of original trial site does not match data of original trial site')
        del original_trial_sites[trialsite_key]
