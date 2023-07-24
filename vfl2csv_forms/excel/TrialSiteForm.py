from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.cell import Cell
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from pandas import ExcelWriter

from vfl2csv_base.TrialSite import TrialSite
from vfl2csv_forms import config
from vfl2csv_forms.excel import styles
from vfl2csv_forms.excel.FormulaColumn import FormulaColumn
from vfl2csv_forms.excel.utilities import EXCEL_COLUMN_NAMES, zero_based_cell_name, zero_based_cell_range_name, \
    zero_based_cell_range, get_character_count
from vfl2csv_forms.excel.utilities import zero_based_cell

dtypes_styles_mapping = {
    'string': styles.table_body_text,
    'Int8': styles.table_body_integer,
    'Int16': styles.table_body_integer,
    'Int32': styles.table_body_integer,
    'Int64': styles.table_body_integer,
    'UInt8': styles.table_body_integer,
    'UInt16': styles.table_body_integer,
    'UInt32': styles.table_body_integer,
    'UInt64': styles.table_body_integer,
    'Float32': styles.table_body_rational,
    'Float64': styles.table_body_rational
}


class TrialSiteForm:
    def __init__(self, trial_site: TrialSite, output_path: Path, formulae_columns: Iterable[FormulaColumn]):
        self.df = trial_site.df
        self.metadata = trial_site.metadata
        self.output_path = output_path
        self.formulae_columns = formulae_columns
        self.sheet_name = trial_site.replace_metadata_keys(config['Output']['excel_sheet_pattern'])
        self.table_head_row = 5
        self.row_span = list(range(self.table_head_row, self.table_head_row + len(self.df.index) + 1))
        self.worksheet = None

        self.conditional_formatting_rules: list[tuple[str, Rule]] = []
        self.add_comment_column()
        # one empty column between dataframe contents and formula columns
        self.first_empty_column = len(self.df.columns) + 1

    # noinspection SpellCheckingInspection
    def add_comment_column(self) -> None:
        # add two new columns:
        # 1. "Aus" for trees that were removed ore are supposed to removed from the forest for various reasons
        # 2. "Bruch" for trees that suffered from a break in the wood
        aus_column_name = EXCEL_COLUMN_NAMES[len(self.df.columns)]
        break_column_name = EXCEL_COLUMN_NAMES[len(self.df.columns) + 1]
        # wrap 'Aus' in space to achieve same string length and consequently column width as 'Bruch'
        self.df.insert(len(self.df.columns), ' Aus ', pd.Series(dtype=pd.StringDtype()))
        self.df.insert(len(self.df.columns), 'Bruch', pd.Series(dtype=pd.StringDtype()))

        if len(self.row_span) < 2:
            # In this case there is only the header column
            return

        # conditional formatting for the entire data row if either Aus or Bruch is not empty
        rule = Rule(
            dxf=DifferentialStyle(fill=PatternFill(bgColor='FDFDD9')),
            type='expression',
            stopIfTrue=True,
            # comma instead of semicolon is required for this to work for some reason
            formula=[f'OR('
                     f'LEN(TRIM(${break_column_name}{self.row_span[1] + 1}))>0,'
                     f'LEN(TRIM(${aus_column_name}{self.row_span[1] + 1}))>0'
                     f')']
        )

        cell_range = f'A{self.row_span[1] + 1}:{break_column_name}{self.row_span[-1] + 1}'

        self.conditional_formatting_rules.append((cell_range, rule))

    def create(self, workbook) -> None:
        worksheet = workbook[self.sheet_name]
        self.worksheet = worksheet
        self.write_metadata()
        self.write_formulae_columns()
        self.apply_table_formatting()
        self.adjust_column_width()
        workbook.save(self.output_path)

    def init_worksheet(self, writer: ExcelWriter) -> None:
        """
        Write the dataframe and return the corresponding workbook and worksheet.
        :return:
        """
        self.df.to_excel(
            writer,
            sheet_name=self.sheet_name,
            startrow=self.table_head_row,
            index=False
        )

    def write_metadata(self) -> None:
        """Insert the metadata into the sheet and format the updated cells."""
        ws = self.worksheet
        # merge ABC and DE for first metadata column, then FG/HI for second metadata column
        metadata_columns = [0, 3, 5, 7]
        for row in range(0, 4):
            self.worksheet.merge_cells(zero_based_cell_range_name(
                metadata_columns[0], row, metadata_columns[0] + 2, row))
            self.worksheet.merge_cells(zero_based_cell_range_name(
                metadata_columns[1], row, metadata_columns[1] + 1, row))
            self.worksheet.merge_cells(zero_based_cell_range_name(
                metadata_columns[2], row, metadata_columns[2] + 1, row))
            self.worksheet.merge_cells(zero_based_cell_range_name(
                metadata_columns[3], row, metadata_columns[3] + 1, row))

        # write the data
        for index, key in enumerate(('Versuch', 'Parzelle', 'Forstamt', 'Revier')):
            zero_based_cell(ws, metadata_columns[0], index).value = f'{key}: '
            zero_based_cell(ws, metadata_columns[1], index).value = self.metadata[key]

        zero_based_cell(ws, metadata_columns[2], 0).value = 'Vermessung am:'
        zero_based_cell(ws, metadata_columns[2], 1).value = 'durch: '

        # apply formatting to the metadata section
        metadata_key_cells = zero_based_cell_range(ws, metadata_columns[0], 0, metadata_columns[0], 3) + \
                             zero_based_cell_range(ws, metadata_columns[2], 0, metadata_columns[2], 3)

        # apply default metadata style only for first column with prefilled data
        metadata_value_cells = zero_based_cell_range(ws, metadata_columns[1], 0, metadata_columns[1], 3)
        for row in metadata_key_cells:
            row[0].style = styles.metadata_keys.name
        for row in metadata_value_cells:
            row[0].style = styles.metadata_values.name

        # for the second metadata value column generated without contents, use separate format
        for row in range(3):
            zero_based_cell(ws, metadata_columns[3], row).style = styles.metadata_values_underlined.name
            zero_based_cell(ws, metadata_columns[3] + 1, row).style = styles.metadata_values_underlined.name

    def write_formulae_columns(self) -> None:
        for formulae_column in self.formulae_columns:
            if formulae_column.yielded_column is not None:
                # in this case, `insert` was already called recursively by another instance
                continue
            self.first_empty_column += formulae_column.insert(self.first_empty_column, self.row_span, self.worksheet)

    def apply_table_formatting(self) -> None:
        """
        Apply most styles and conditional formatting rules, excluding the metadata keys/values as well as table values
        (except for headers) not part of the dataframe, that is formula columns and comment columns (Aus/Bruch).
        """
        for column_index in range(self.first_empty_column):
            # apply style to header of all columns
            zero_based_cell(self.worksheet, column_index, self.table_head_row).style = styles.table_head.name
            if column_index < len(self.df.columns):
                # Do only format the values below the header if the column index is still contained in the dataframe
                # Custom columns that are not contained in the dataframe should be formatted separately.
                style = dtypes_styles_mapping.get(
                    self.df.dtypes[column_index].name,
                    styles.table_body_text
                ).name
                for row in self.row_span[1:]:
                    zero_based_cell(self.worksheet, column_index, row).style = style
            elif column_index == len(self.df.columns):
                # First column after dataframe contents that is emptys
                for row in self.row_span[1:]:
                    zero_based_cell(self.worksheet, column_index, row).style = styles.table_body_text

        # apply conditional formatting rules
        for key, rule in self.conditional_formatting_rules:
            self.worksheet.conditional_formatting.add(key, rule)

        # freeze the header rows and no columns
        self.worksheet.freeze_panes = zero_based_cell_name(0, self.table_head_row + 1)

    def adjust_column_width(self) -> None:
        for column in range(self.first_empty_column):
            column_content: list[list[Cell]] = zero_based_cell_range(self.worksheet, column, self.row_span[0], column,
                                                                     self.row_span[-1])
            max_length = 0
            if column_content[0][0].value is None:
                continue
            else:
                max_length = get_character_count(column_content[0][0].value, decimal_digits=1)

            # only scan content if there is any content
            if len(column_content) >= 2:
                first_content_value: str = column_content[1][0].value
                if not (isinstance(first_content_value, str) and first_content_value.startswith('=')):
                    # If condition above is not true, then we have a column with formulas and cannot compute the width
                    # of the formula's output
                    for row in column_content:
                        max_length = max(max_length, get_character_count(row[0].value, decimal_digits=1))

            # adding one and multiplying by 1.3 fits well
            column_width = 1.3 * (max_length + 1)
            self.worksheet.column_dimensions[get_column_letter(column + 1)].width = column_width
