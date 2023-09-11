from openpyxl.formatting.rule import CellIsRule, Rule, ColorScale, FormatObject
from openpyxl.styles import NamedStyle, Font, Alignment, Color, Border, Side
from openpyxl.workbook import Workbook

_full_cell_border = Border(
    bottom=Side(style="thin"),
    left=Side(style="thin"),
    top=Side(style="thin"),
    right=Side(style="thin"),
)

metadata_keys = NamedStyle(
    name="metadata key cells",
    font=Font(bold=True, name="Calibri", size=12),
    alignment=Alignment(horizontal="right"),
)

metadata_values = NamedStyle(
    name="metadata value cells", font=Font(name="Calibri", size=12)
)

metadata_values_underlined = NamedStyle(
    name="underlined metadata value cells",
    font=Font(name="Calibri", size=12),
    border=Border(bottom=Side(style="thin")),
)

table_head = NamedStyle(
    name="table head cells",
    font=Font(bold=True, name="Calibri", size=12),
    alignment=Alignment(horizontal="center", vertical="center"),
)
table_head_border = NamedStyle(
    name="table head cells (with outlines)",
    font=Font(bold=True, name="Calibri", size=12),
    alignment=Alignment(horizontal="center", vertical="center"),
    border=_full_cell_border,
)

table_body_text = NamedStyle(
    name="table body cells (text)", font=Font(name="Calibri", size=12)
)
table_body_integer = NamedStyle(
    name="table body cells (integer)",
    font=Font(name="Calibri", size=12),
    number_format="0",
)
table_body_rational = NamedStyle(
    name="table body cells (rational)",
    font=Font(name="Calibri", size=12),
    number_format="0.0",
)

table_body_text_border = NamedStyle(
    name="table body cells (text, with outlines)",
    font=Font(name="Calibri", size=12),
    border=_full_cell_border,
)
table_body_integer_border = NamedStyle(
    name="table body cells (integer, with outlines)",
    font=Font(name="Calibri", size=12),
    number_format="0",
    border=_full_cell_border,
)
table_body_rational_border = NamedStyle(
    name="table body cells (rational, with outlines)",
    font=Font(name="Calibri", size=12),
    number_format="0.0",
    border=_full_cell_border,
)


# noinspection SpellCheckingInspection,PyDefaultArgument
def full_conditional_formatting_list(__mutable={"count": 1}) -> list[Rule]:
    # __mutable['count'] counts the calls to this function, incrementing by one after every call.
    # This is required so returned conditional formatting rules have different priorities all the time, while still
    # maintaining the difference between priorities of the same rules.
    # This is required to fix a bug, see https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1941

    # alternatively to the color scale defined below: set background to one of three colors depending on whether the
    # number is positive, negative or zero
    """
    conditional_formatting_greater_than = CellIsRule(
        operator='greaterThan',
        formula=['0'],
        stopIfTrue=True,
        fill=PatternFill(bgColor='ccffcc')
    )
    conditional_formatting_greater_than.priority = __mutable['count'] + 1000
    conditional_formatting_equal = CellIsRule(
        operator='equal',
        formula=['0'],
        stopIfTrue=True,
        fill=PatternFill(bgColor='FFEB9C')
    )
    conditional_formatting_equal.priority = __mutable['count'] + 1000
    conditional_formatting_less_than = CellIsRule(
        operator='lessThan',
        formula=['0'],
        stopIfTrue=True,
        fill=PatternFill(bgColor='ffcccc')
    )
    conditional_formatting_less_than.priority = __mutable['count'] + 1000
    """

    color_scale = Rule(
        type="colorScale",
        colorScale=ColorScale(
            cfvo=[
                FormatObject(type="min"),
                FormatObject(type="num", val=0),
                FormatObject(type="max"),
            ],
            color=[
                # use a very aggressive red color combined with a pale green tone so that negative values stand out
                # strongly
                Color("C00000"),
                Color("FFFFFF"),
                Color("63BE7B"),
            ],
        ),
        priority=1000,
    )

    # in Excel, the formula `= "" > 0` returns true. To prevent wrong formatting on empty columns, we therefore need to
    # cancel applying formatting if the cell is an empty string.
    conditional_formatting_empty = CellIsRule(
        operator="equal",
        formula=['""'],
        stopIfTrue=True,
        # fill=PatternFill(bgColor='E7E6E6')
    )
    conditional_formatting_empty.priority = __mutable["count"]
    __mutable["count"] += 1
    return [conditional_formatting_empty, color_scale]


def register(workbook: Workbook) -> None:
    for style in (
            metadata_keys,
            metadata_values,
            metadata_values_underlined,
            table_head,
            table_head_border,
            table_body_text,
            table_body_integer,
            table_body_rational,
            table_body_text_border,
            table_body_integer_border,
            table_body_rational_border,
    ):
        workbook.add_named_style(style)
