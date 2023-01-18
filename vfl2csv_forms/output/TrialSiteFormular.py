from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from pandas import ExcelWriter

from vfl2csv_base.TrialSite import TrialSite
from vfl2csv_forms import config
from vfl2csv_forms.excel import styles
from vfl2csv_forms.excel.utilities import EXCEL_COLUMN_NAMES
from vfl2csv_forms.excel.utilities import zeroBasedCell
from vfl2csv_forms.output.FormulaColumn import FormulaColumn

dtypes_styles_mapping = {
    'string': styles.table_body_text,
    'Int8': styles.table_body_integer,
    'Int16': styles.table_body_integer,
    'Int32': styles.table_body_integer,
    'Int64': styles.table_body_integer,
    'UInt8': styles.table_body_integer,
    'UInt16': styles.table_body_integer,
    'UInt32': styles.table_body_integer,
    'UInt64': styles.table_body_integer,
    'Float32': styles.table_body_rational,
    'Float64': styles.table_body_rational
}


class TrialSiteFormular:
    def __init__(self, trial_site: TrialSite, output_path: Path, formulae_columns: Iterable[FormulaColumn]):
        self.df = trial_site.df
        self.metadata = trial_site.metadata
        self.output_path = output_path
        self.formulae_columns = formulae_columns
        self.sheet_name = trial_site.replace_metadata_keys(config['Output']['excel_sheet_pattern'])
        self.table_head_row = 5
        self.row_span = list(range(self.table_head_row, self.table_head_row + len(self.df.index) + 1))
        self.worksheet = None

        self.conditional_formatting_rules: list[tuple[str, Rule]] = []
        self.add_comment_column()
        self.first_empty_column = len(self.df.columns) + 1

    def add_comment_column(self):
        # add two new columns:
        # 1. "Aus" for trees that were removed ore are supposed to removed from the forest for various reasons
        # 2. "Bruch" for trees that suffered from a break in the wood
        aus_column_name = EXCEL_COLUMN_NAMES[len(self.df.columns)]
        break_column_name = EXCEL_COLUMN_NAMES[len(self.df.columns) + 1]
        # wrap 'Aus' in space to achieve same string length and consequently column width as 'Bruch'
        self.df.insert(len(self.df.columns), ' Aus ', pd.Series(dtype=pd.StringDtype()))
        self.df.insert(len(self.df.columns), 'Bruch', pd.Series(dtype=pd.StringDtype()))
        # conditional formatting for the entire data row if either Aus or Bruch is not empty
        rule = Rule(
            dxf=DifferentialStyle(fill=PatternFill(bgColor='FDFDD9')),
            type='expression',
            stopIfTrue=True,
            # comma instead of semicolon is required for this to work for some reason
            formula=[f'OR('
                     f'LEN(TRIM(${break_column_name}{self.row_span[1] + 1}))>0,'
                     f'LEN(TRIM(${aus_column_name}{self.row_span[1] + 1}))>0'
                     f')']
        )

        cell_range = f'A{self.row_span[1] + 1}:{break_column_name}{self.row_span[-1] + 1}'

        self.conditional_formatting_rules.append((cell_range, rule))

    def create(self, workbook):
        # workbook = load_workbook(self.output_path)
        worksheet = workbook[self.sheet_name]
        self.worksheet = worksheet
        self.write_metadata()
        self.write_formulae_columns()
        self.apply_formatting()
        self.adjust_column_width()
        workbook.save(self.output_path)

    def init_worksheet(self, writer: ExcelWriter):
        """
        Write the dataframe and return the corresponding workbook and worksheet.
        :return:
        """
        self.df.to_excel(
            writer,
            sheet_name=self.sheet_name,
            startrow=self.table_head_row,
            index=False
        )

    def write_metadata(self):
        # merge two horizontally adjacent cells each
        for column in ('A', 'C', 'F', 'H'):
            letters = (column, get_column_letter(column_index_from_string(column) + 1))
            for row in range(1, 5):
                self.worksheet.merge_cells(f'{letters[0]}{row}:{letters[1]}{row}')

        # write the data
        for index, key in enumerate(('Versuch', 'Parzelle', 'Forstamt', 'Revier')):
            self.worksheet[f'A{index + 1}'] = f'{key}: '
            self.worksheet[f'C{index + 1}'] = self.metadata[key]
        self.worksheet['F1'] = 'Vermessung am: '
        self.worksheet['F2'] = 'durch: '

    def write_formulae_columns(self):
        for formulae_column in self.formulae_columns:
            if formulae_column.yielded_column is not None:
                # in this case, `insert` was already called recursively by another instance
                continue
            self.first_empty_column += formulae_column.insert(self.first_empty_column, self.row_span, self.worksheet)

    def apply_formatting(self):
        for row in self.worksheet['A1:A4'] + self.worksheet['F1:F4']:
            row[0].style = styles.metadata_keys.name
        for row in self.worksheet['C1:C4'] + self.worksheet['H1:H4']:
            row[0].style = styles.metadata_values.name

        for column_index in range(self.first_empty_column):
            # header column
            zeroBasedCell(self.worksheet, self.table_head_row, column_index).style = styles.table_head.name
            if len(self.df.columns) > column_index:
                style = dtypes_styles_mapping.get(
                    self.df.dtypes[column_index].name,
                    styles.table_body_text
                ).name
                for row in self.row_span[1:]:
                    zeroBasedCell(self.worksheet, row, column_index).style = style

        for key, rule in self.conditional_formatting_rules:
            self.worksheet.conditional_formatting.add(key, rule)

    def adjust_column_width(self):
        for column in range(self.first_empty_column):
            table_head_cell = zeroBasedCell(self.worksheet, self.table_head_row, column)
            # factor 1.5 is an arbitrary value that fits quite nicely
            if table_head_cell.value is not None:
                self.worksheet.column_dimensions[get_column_letter(column + 1)].width = 1.5 * len(table_head_cell.value)
